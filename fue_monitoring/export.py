# SPDX-License-Identifier: AGPL-3.0-or-later
# Copyright (C) 2026 Tobias Hardes

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import FundingCall

COLUMNS = [
    ("source", "Source"),
    ("title", "Title"),
    ("programme", "Programme"),
    ("funding_rate", "Funding Rate"),
    ("deadline", "Deadline"),
    ("deadline_status", "Deadline Status"),
    ("budget", "Budget"),
    ("matched_keywords", "Matched Keywords"),
    ("note", "Note"),
    ("link", "Link"),
]

STATUS_FILL = {
    "expired": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
    "urgent": PatternFill(start_color="FCE8B2", end_color="FCE8B2", fill_type="solid"),
    "open": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
    "unknown": PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid"),
}


def write_excel(entries: list[FundingCall], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Funding Calls"

    for col_idx, (_, header) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    for row_idx, entry in enumerate(entries, start=2):
        for col_idx, (field_name, _) in enumerate(COLUMNS, start=1):
            value = getattr(entry, field_name)
            if field_name == "matched_keywords":
                value = ", ".join(value)
            elif field_name == "deadline" and value is not None:
                value = value.strftime("%d.%m.%Y")
            ws.cell(row=row_idx, column=col_idx, value=value)

        fill = STATUS_FILL.get(entry.deadline_status)
        if fill:
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(entries) + 1, 1)}"
    ws.freeze_panes = "A2"

    widths = [18, 55, 20, 14, 12, 13, 14, 25, 30, 45]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
